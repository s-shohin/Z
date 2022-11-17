#!/usr/bin/env python
# coding: utf-8

# In[1]:


import Z_func as Z
from tqdm import tqdm
from time import sleep

import pandas as pd
import openpyxl as xl

import os
import sys
import random
import subprocess


# In[ ]:


FILE_NAME=input('ファイル名を入力（拡張子.xlsmは除く）=')
SHEET_NAME='Z打鍵'
BASE_BAT_SIZE=int(input('バッチあたりの打鍵数(1=<) = ')) #バッチサイズ。一度の打鍵件数

#読み込み
wb=xl.load_workbook(FILE_NAME+'.xlsm',keep_vba=True)    
df=pd.read_excel(FILE_NAME+'.xlsm',sheet_name=SHEET_NAME)

#列を追加
df['車両AMTエラー']=''
df['新車保険金額エラー']=''

ws=wb[SHEET_NAME]
ws.cell(row=1, column=df.columns.get_loc('車両AMTエラー')+1).value = '車両AMTエラー'
ws.cell(row=1, column=df.columns.get_loc('新車保険金額エラー')+1).value = '新車保険金額エラー'

#打鍵する行のリストの箱を用意
calc_row=list(range(1))

#打鍵、バッチ単位のループ
while len(calc_row) > 0:
    #打鍵する行を特定
    calc_row=list()
    #数値のはいっていない行のうち、バッチサイズの行だけExcel上の行番号を取得（dfのindex+2）する。
    BAT_SIZE=BASE_BAT_SIZE + random.randint(0, BASE_BAT_SIZE) #並列処理時にデータファイルアクセスのタイミングをずらすために乱数を加算
    calc_row=list(df[df['車有P'].isna()].index[0:BAT_SIZE]+2) #Eは再計算しない。打鍵条件で通勤＋15000km?など選択不可なものもあるため。
    print(calc_row)

    #エラー回数カウントを初期化
    error_count = 0

    #打鍵を始める行に打鍵中と入力
    for j in calc_row:
        ws.cell(row=j, column=df.columns.get_loc('車有P')+1).value = '打鍵中' 
        df.iloc[j-2,df.columns.get_loc('車有P')]='打鍵中'
    wb.save(FILE_NAME+'.xlsm')#いったん保存

    #####打鍵、行単位のループ#####################################  
    for i in tqdm(calc_row):
        data=df.loc[i-2,:].to_dict()
        data = Z.Z_func(data)#打鍵

        #結果をdfに書き込む
        df_temp=pd.DataFrame.from_dict(data, orient='index').T
        df.loc[i-2,:] = df_temp.iloc[0,:]
        if data['車有P'] == 'E':
            error_count = error_count + 1
    #####行単位のループ終了######################################

    if error_count == BAT_SIZE:
        print('1時間停止中')
        sleep(3600)#すべてE、エラーだったら、たぶんHPメンテ中と判断して、1時間停止

    ########並列で実行するため、あらためて現時点の最新版のファイルを読み出して結果を追加

    wb=xl.load_workbook(FILE_NAME+'.xlsm',keep_vba=True)

    ws=wb[SHEET_NAME]
    for i in calc_row:
        ws.cell(row=i, column=73).value = df.loc[i-2,'車有P'] 
        ws.cell(row=i, column=74).value = df.loc[i-2,'車無P']
        ws.cell(row=i, column=75).value = df.loc[i-2,'イ割車有'] 
        ws.cell(row=i, column=76).value = df.loc[i-2,'イ割車無']
        ws.cell(row=i, column=77).value = df.loc[i-2,'早割'] 
        ws.cell(row=i, column=78).value = df.loc[i-2,'車両AMTエラー']
        ws.cell(row=i, column=79).value = df.loc[i-2,'新車保険金額エラー']


    #いったん保存
    wb.save(FILE_NAME+'.xlsm')
    #dfを更新
    df=pd.read_excel(FILE_NAME+'.xlsm',sheet_name=SHEET_NAME)


#バッチ単位のループ終了########################################

    

